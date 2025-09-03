# src/main.py

import argparse
import json
import os
import smtplib
from datetime import datetime
from email.mime.text import MIMEText

import pandas as pd
from openpyxl.utils import get_column_letter

from src.scrapers.greenhouse import scrape_greenhouse_jobs
from src.scrapers.ashby import scrape_ashby_jobs
from src.scrapers.lever import scrape_lever_jobs
from src.scrapers.custom.google import scrape_google_jobs
from src.scrapers.custom.shopify import scrape_shopify_jobs
from src.scrapers.custom.microsoft import scrape_microsoft_jobs
from src.scrapers.custom.meta import scrape_meta_jobs
from src.scrapers.custom.apple import scrape_apple_jobs
from src.scrapers.custom.amazon import scrape_amazon_jobs
from src.scrapers.custom.workday import scrape_workday_jobs

# Optional: self-healing resolver (kept for future use, but we now prefer explicit config first)
from src.selfheal import resolve_company_source

# ------------------------------------------------
# Configuration
# ------------------------------------------------
CONFIG_PATH = "config/companies.json"
OUTPUT_DIR = "data/raw/"
STATE_FILE = "data/seen_jobs.txt"
KEYWORD_FILTERS = [
    "new grad",
    "university grad",
    "entry-level",
    "entry level",
    "software engineer",
    "software engineer i",
    "swe i",
    "step",
    "residency",
    "early career",
    "futureforce",
    "class of",
    "extreme blue",
    "aspire",
    "launchpad",
    "rotational",
    "graduate program",
    "associate",
]

NETWORKING_HEADERS = [
    "Company",
    "Relevant Job Example",
    "Entry Level SE Search",
    "General Role Search",
    "Status",
    "Notes",
]

# ------------------------------------------------
# Helpers
# ------------------------------------------------
def _parse_csv_flag(val: str | None) -> set[str]:
    if not val:
        return set()
    return {s.strip().lower() for s in val.split(",") if s.strip()}

def load_seen_jobs() -> set[str]:
    try:
        with open(STATE_FILE, "r") as f:
            return set(line.strip() for line in f if line.strip())
    except FileNotFoundError:
        return set()

def save_seen_jobs(job_urls: set[str]) -> None:
    os.makedirs(os.path.dirname(STATE_FILE), exist_ok=True)
    with open(STATE_FILE, "w") as f:
        for url in sorted(job_urls):
            f.write(f"{url}\n")

def send_email_notification(new_jobs_df: pd.DataFrame) -> None:
    sender_email = os.environ.get("SENDER_EMAIL")
    sender_password = os.environ.get("SENDER_PASSWORD")
    if not sender_email or not sender_password:
        print("  > Email credentials not set. Skipping email notification.")
        return

    subject = f"Found {len(new_jobs_df)} New Job Postings!"
    body_lines = ["Here are the new jobs found today", ""]
    for _, job in new_jobs_df.iterrows():
        body_lines.append(f"- {job.get('Title','')} at {job.get('Company','')}")
        body_lines.append(f"  Location: {job.get('Location','')}")
        body_lines.append(f"  Link: {job.get('URL','')}")
        if "Posted" in job and job.get("Posted"):
            body_lines.append(f"  Posted: {job.get('Posted')}")
        body_lines.append("")
    body = "\n".join(body_lines)

    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = sender_email
    msg["To"] = sender_email

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(sender_email, sender_password)
            s.sendmail(sender_email, sender_email, msg.as_string())
        print("  > Email notification sent successfully!")
    except Exception as e:
        print(f"  > Failed to send email. Error: {e}")

def add_hyperlinks_to_column(ws, header_text: str) -> None:
    header_row = 1
    col_idx = None
    for c in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row, column=c).value
        if cell_val is not None and str(cell_val).strip() == header_text:
            col_idx = c
            break
    if not col_idx:
        return

    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        url = str(cell.value).strip() if cell.value else ""
        if url.startswith("http"):
            cell.hyperlink = url
            cell.style = "Hyperlink"

def autosize_columns(ws, min_width: int = 8, max_width: int = 60) -> None:
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            v = str(v) if v is not None else ""
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))

def build_networking_rows(jobs: list[dict]) -> list[dict]:
    by_company = {}
    for j in jobs:
        comp = str(j.get("Company", "")).strip()
        if not comp:
            continue
        if comp not in by_company:
            by_company[comp] = j

    rows = []
    for comp, j in sorted(by_company.items(), key=lambda kv: kv[0].lower()):
        rows.append({
            "Company": comp,
            "Title": j.get("Title", ""),
            "URL": j.get("URL", ""),
            "Entry_Level_SE_Search": j.get("Entry_Level_SE_Search", ""),
            "General_Role_Search": j.get("General_Role_Search", ""),
            "Status": "",
            "Notes": "",
        })
    return rows

def write_networking_sheet(writer, jobs: list[dict]) -> None:
    wb = writer.book
    ws = wb.create_sheet("Networking To-Do")
    ws.append(NETWORKING_HEADERS)

    rows = build_networking_rows(jobs)
    for r in rows:
        ws.append([
            r["Company"],
            r["Title"],                 # clickable via hyperlink below
            r["Entry_Level_SE_Search"], # must be populated by scrapers
            r["General_Role_Search"],   # must be populated by scrapers
            r["Status"],
            r["Notes"],
        ])

    # Map data rows to their job URLs for hyperlinking the title
    url_map = {i + 2: rows[i]["URL"] for i in range(len(rows))}
    header_cells = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    title_col = header_cells.get("Relevant Job Example")
    se_col = header_cells.get("Entry Level SE Search")
    role_col = header_cells.get("General Role Search")

    if title_col:
        for row_idx, url in url_map.items():
            if url and str(url).startswith("http"):
                cell = ws.cell(row=row_idx, column=title_col)
                cell.hyperlink = url
                cell.style = "Hyperlink"

    if se_col:
        for row_idx in range(2, ws.max_row + 1):
            url = ws.cell(row=row_idx, column=se_col).value
            if url and str(url).startswith("http"):
                c = ws.cell(row=row_idx, column=se_col)
                c.hyperlink = url
                c.style = "Hyperlink"

    if role_col:
        for row_idx in range(2, ws.max_row + 1):
            url = ws.cell(row=row_idx, column=role_col).value
            if url and str(url).startswith("http"):
                c = ws.cell(row=row_idx, column=role_col)
                c.hyperlink = url
                c.style = "Hyperlink"

    autosize_columns(ws)

# ------------------------------------------------
# Target gating
# ------------------------------------------------
ALL_TARGETS = {
    "greenhouse",
    "ashby",
    "lever",
    "google",
    "shopify",
    "microsoft",
    "meta",
    "apple",
    "amazon",
    "workday",
}

def _should_run(name: str, only: set[str], skip: set[str]) -> bool:
    if name in skip:
        return False
    if not only or "all" in only:
        return True
    return name in only

# ------------------------------------------------
# Main run
# ------------------------------------------------
def run_scrapers(only: set[str] | None = None, skip: set[str] | None = None) -> None:
    """
    Run the scrapers. 'only' and 'skip' are sets of target names (lowercase).
    Defaults let this be called without arguments (used by tests).
    """
    only = only or set()
    skip = skip or set()

    seen_job_urls = load_seen_jobs()
    print(f"Loaded {len(seen_job_urls)} previously seen jobs from state file.")

    # Load config
    try:
        with open(CONFIG_PATH, "r") as f:
            companies_config = json.load(f)
    except FileNotFoundError:
        print(f"Error: Configuration file not found at '{CONFIG_PATH}'")
        return

    greenhouse_companies = companies_config.get("greenhouse", {}) or {}
    ashby_companies      = companies_config.get("ashby", {}) or {}
    lever_companies      = companies_config.get("lever", {}) or {}
    workday_portals      = companies_config.get("workday", {}) or {}

    all_current_jobs: list[dict] = []

    # 1) Run explicit config first (no resolver required — great for tests)
    if _should_run("greenhouse", only, skip) and greenhouse_companies:
        print("--- Starting Greenhouse Scrape ---")
        for company, token in greenhouse_companies.items():
            try:
                jobs = scrape_greenhouse_jobs(company, token, KEYWORD_FILTERS)
            except Exception as e:
                print(f"  > {company} (Greenhouse) failed: {e}")
                jobs = []
            all_current_jobs.extend(jobs)

    if _should_run("ashby", only, skip) and ashby_companies:
        print("\n--- Starting Ashby Scrape ---")
        for company, token in ashby_companies.items():
            try:
                jobs = scrape_ashby_jobs(company, token, KEYWORD_FILTERS)
            except Exception as e:
                print(f"  > {company} (Ashby) failed: {e}")
                jobs = []
            all_current_jobs.extend(jobs)

    if _should_run("lever", only, skip) and lever_companies:
        print("\n--- Starting Lever Scrape ---")
        for company, token in lever_companies.items():
            try:
                jobs = scrape_lever_jobs(company, token, KEYWORD_FILTERS)
            except Exception as e:
                print(f"  > {company} (Lever) failed: {e}")
                jobs = []
            all_current_jobs.extend(jobs)

    # 2) Optional self-healing pass for companies that exist in config but had no jobs
    #    Useful outside of tests; safe no-op in tests.
    #    (We try to detect ATS if explicit config produced nothing.)
    if (_should_run("greenhouse", only, skip) or _should_run("ashby", only, skip)) and not all_current_jobs:
        print("\n--- Self-healing ATS detection (fallback) ---")
        companies = sorted(set(greenhouse_companies) | set(ashby_companies))
        for company in companies:
            candidate_slugs = []
            for s in (greenhouse_companies.get(company), ashby_companies.get(company)):
                if s and s not in candidate_slugs:
                    candidate_slugs.append(s)
            if not candidate_slugs:
                candidate_slugs = [company]

            ats = real_slug = None
            for cand in candidate_slugs:
                try:
                    ats, real_slug = resolve_company_source(company, cand)
                except Exception as e:
                    print(f"  > Resolver error for {company}/{cand}: {e}")
                    ats = real_slug = None
                if ats:
                    break

            try:
                if ats == "greenhouse":
                    jobs = scrape_greenhouse_jobs(company, real_slug, KEYWORD_FILTERS)
                elif ats == "ashby":
                    jobs = scrape_ashby_jobs(company, real_slug, KEYWORD_FILTERS)
                elif ats == "lever":
                    jobs = scrape_lever_jobs(company, real_slug, KEYWORD_FILTERS)
                else:
                    print(f"  > {company}: No valid ATS found (skipping)")
                    jobs = []
            except Exception as e:
                print(f"  > {company} (self-heal) failed: {e}")
                jobs = []

            all_current_jobs.extend(jobs)

    # 3) Workday (config-driven)
    if _should_run("workday", only, skip) and workday_portals:
        print("\n--- Starting Workday Scrape ---")
        try:
            wd_jobs = scrape_workday_jobs(KEYWORD_FILTERS, workday_portals)
        except Exception as e:
            print(f"  > Workday scrape failed: {e}")
            wd_jobs = []
        all_current_jobs.extend(wd_jobs)

    # 4) Custom scrapes (guarded with try/except so tests don't blow up)
    print("\n--- Starting Custom Scrapes ---")
    if _should_run("google", only, skip):
        try:
            all_current_jobs.extend(scrape_google_jobs(KEYWORD_FILTERS))
        except Exception as e:
            print(f"  > Google scrape failed: {e}")
    if _should_run("shopify", only, skip):
        try:
            all_current_jobs.extend(scrape_shopify_jobs(KEYWORD_FILTERS))
        except Exception as e:
            print(f"  > Shopify scrape failed: {e}")
    if _should_run("microsoft", only, skip):
        try:
            all_current_jobs.extend(scrape_microsoft_jobs(KEYWORD_FILTERS))
        except Exception as e:
            print(f"  > Microsoft scrape failed: {e}")
    if _should_run("meta", only, skip):
        try:
            all_current_jobs.extend(scrape_meta_jobs(KEYWORD_FILTERS))
        except Exception as e:
            print(f"  > Meta scrape failed: {e}")
    if _should_run("apple", only, skip):
        try:
            all_current_jobs.extend(scrape_apple_jobs(KEYWORD_FILTERS))
        except Exception as e:
            print(f"  > Apple scrape failed: {e}")
    if _should_run("amazon", only, skip):
        try:
            all_current_jobs.extend(scrape_amazon_jobs(KEYWORD_FILTERS))
        except Exception as e:
            print(f"  > Amazon scrape failed: {e}")

    # Nothing found
    if not all_current_jobs:
        print("\nNo jobs found at all during this run.")
        save_seen_jobs(seen_job_urls)
        return

    # De-dupe against seen set using URL
    new_jobs_found: list[dict] = []
    current_job_urls: set[str] = set()
    for job in all_current_jobs:
        url = (job.get("URL") or "").strip()
        if not url:
            continue
        current_job_urls.add(url)
        if url not in seen_job_urls:
            # Normalize LinkedIn field names if some scrapers use different keys
            if "Alumni_Search_URL" in job and "Entry_Level_SE_Search" not in job:
                job["Entry_Level_SE_Search"] = job.get("Alumni_Search_URL")
            if "Role_Search_URL" in job and "General_Role_Search" not in job:
                job["General_Role_Search"] = job.get("Role_Search_URL")
            new_jobs_found.append(job)

    print(f"\nScraping complete. Found {len(new_jobs_found)} new jobs.")

    if new_jobs_found:
        new_jobs_df = pd.DataFrame(new_jobs_found)

        # Email
        send_email_notification(new_jobs_df)

        # Write Excel
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        today_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"{OUTPUT_DIR}NEW_jobs_report_{today_str}.xlsx"
        print(f"  > Saving report to multi-sheet Excel file: '{filename}'")

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # Networking sheet first
            write_networking_sheet(writer, new_jobs_found)

            # Company sheets (include 'Posted' and sort if present)
            for company_name, company_df in new_jobs_df.groupby("Company"):
                print(f"    - Writing sheet for {company_name}...")
                df = company_df.copy()
                if "Posted" in df.columns:
                    df["_Posted_dt"] = pd.to_datetime(df["Posted"], errors="coerce")
                    df = df.sort_values("_Posted_dt", ascending=False, na_position="last").drop(columns=["_Posted_dt"])
                cols = ["Company", "Title", "URL", "Location", "Posted",
                        "Entry_Level_SE_Search", "General_Role_Search"]
                safe_cols = [c for c in cols if c in df.columns]
                df = df[safe_cols]
                df.to_excel(writer, sheet_name=company_name, index=False)
                ws = writer.sheets[company_name]
                add_hyperlinks_to_column(ws, "URL")
                add_hyperlinks_to_column(ws, "Entry Level SE Search")
                add_hyperlinks_to_column(ws, "General Role Search")
                autosize_columns(ws)

        print("  > Multi-sheet Excel file with Networking To-Do list saved successfully.")

    # Update state
    save_seen_jobs(current_job_urls)
    print(f"Updated state file with {len(current_job_urls)} current jobs for next run.")

def _cli() -> tuple[set[str], set[str]]:
    parser = argparse.ArgumentParser(description="Run job scrapers")
    parser.add_argument("--only", type=str, default=os.getenv("ONLY", ""), help="Comma separated targets to run. Example: --only greenhouse,ashby or --only all")
    parser.add_argument("--skip", type=str, default=os.getenv("SKIP", ""), help="Comma separated targets to skip. Example: --skip apple,amazon")
    args = parser.parse_args()
    only = _parse_csv_flag(args.only)
    skip = _parse_csv_flag(args.skip)

    # validate names
    unknown = (only | skip) - (ALL_TARGETS | {"all"})
    if unknown:
        print(f"Warning: unknown targets ignored: {', '.join(sorted(unknown))}")
        only -= unknown
        skip -= unknown
    return only, skip

if __name__ == "__main__":
    only, skip = _cli()
    run_scrapers(only, skip)